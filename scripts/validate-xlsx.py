"""Independent validation of a generated .xlsx.

Deliberately uses libraries that know nothing about how the file was written:
`zipfile` for the container and `openpyxl` for the workbook. Checks both the
spreadsheet content (styles, values, row count) and the ZIP-level properties
an OOXML consumer expects — compression, ZIP version 2.0, no ZIP64.

Usage: python3 scripts/validate-xlsx.py out/example-node.xlsx [expected_rows]
"""

import struct
import sys
import zipfile

import openpyxl

ZIP64_MARKER = 45  # "version needed to extract" value that means ZIP64
FLAG_DATA_DESCRIPTOR = 0x0008


def check_container(path):
    """Checks the ZIP container itself, below the spreadsheet level."""
    with zipfile.ZipFile(path) as archive:
        assert archive.testzip() is None, "corrupt entry in the archive"
        names = archive.namelist()

        for info in archive.infolist():
            assert info.create_version <= 20, (
                f"{info.filename}: created with ZIP {info.create_version / 10}, "
                "Office requires 2.0"
            )
            assert info.extract_version <= 20, (
                f"{info.filename}: needs ZIP {info.extract_version / 10} to extract, "
                "Office requires 2.0"
            )
            assert info.extract_version < ZIP64_MARKER, f"{info.filename}: ZIP64 entry"

        with open(path, "rb") as handle:
            raw = handle.read()
        assert raw.find(b"PK\x06\x06") == -1, "ZIP64 end of central directory present"

        sheet = archive.getinfo("xl/worksheets/sheet1.xml")
        return {
            "names": names,
            "sheet_method": sheet.compress_type,
            "sheet_raw": sheet.file_size,
            "sheet_stored": sheet.compress_size,
            "streamed": bool(sheet.flag_bits & FLAG_DATA_DESCRIPTOR),
        }


def check_local_header_is_streamed(path):
    """The local header must declare unknown sizes, which is what lets the
    archive be written before the row count is known."""
    with open(path, "rb") as handle:
        raw = handle.read(4096)
    assert raw[:4] == b"PK\x03\x04", "missing local file header signature"
    flags, method, _, crc, csize, usize = struct.unpack("<HHIIII", raw[6:26])
    assert flags & FLAG_DATA_DESCRIPTOR, "sizes were known upfront (not streamed)"
    assert (crc, csize, usize) == (0, 0, 0), "local header carries sizes despite bit 3"
    return method


def check_workbook(path, expected_rows):
    book = openpyxl.load_workbook(path)
    sheet = book.active

    header = [cell.value for cell in sheet[1]]
    assert all(sheet.cell(row=1, column=i + 1).font.bold for i in range(len(header))), (
        "header row is not bold"
    )

    pk = sheet.cell(row=1, column=1)
    assert pk.fill.fgColor.rgb == "FFFFE699", f"PK header fill is {pk.fill.fgColor.rgb}"
    assert sheet.cell(row=2, column=1).fill.fgColor.rgb == "FFFFE699", "PK cell not filled"
    assert sheet.cell(row=2, column=2).fill.fill_type in (None, "none"), (
        "non-PK cell should be unstyled"
    )
    assert not sheet.cell(row=2, column=2).font.bold, "data row should not be bold"

    if expected_rows is not None:
        assert sheet.max_row == expected_rows, (
            f"expected {expected_rows} rows, found {sheet.max_row}"
        )

    return sheet.max_row, header, sheet.cell(row=sheet.max_row, column=2).value


def main():
    path = sys.argv[1]
    expected_rows = int(sys.argv[2]) if len(sys.argv) > 2 else None

    container = check_container(path)
    header_method = check_local_header_is_streamed(path)
    rows, header, last_name = check_workbook(path, expected_rows)

    assert container["sheet_method"] == zipfile.ZIP_DEFLATED, "worksheet is not deflated"
    assert header_method == zipfile.ZIP_DEFLATED, "local header does not declare deflate"
    assert container["streamed"], "worksheet was not written with a data descriptor"

    ratio = container["sheet_stored"] / container["sheet_raw"] * 100
    print(f"OK  {path}")
    print(f"    parts        {len(container['names'])}: {', '.join(container['names'])}")
    print(f"    rows         {rows} (header + {rows - 1} data), last name {last_name!r}")
    print(f"    columns      {header}")
    print("    zip          2.0, no ZIP64, deflate, streamed (data descriptor)")
    print(
        f"    sheet1.xml   {container['sheet_raw']:,} B -> "
        f"{container['sheet_stored']:,} B ({ratio:.1f}%)"
    )


if __name__ == "__main__":
    main()
